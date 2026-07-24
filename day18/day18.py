import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CREATE DATASETS
# ==========================================
# Sales transactions
data = {
    "Region": ["North", "South", "East", "West", "North", "South", "East", "West"],
    "Rep": ["Alice", "Bob", "Charlie", "Diana", "Alice", "Bob", "Charlie", "Diana"],
    "Q1_Sales": [15000, 22000, 18000, 12000, 16000, 24000, 19000, 11000],
    "Q2_Sales": [18000, 21000, 19500, 14000, 17500, 26000, 21000, 13000],
}

df = pd.DataFrame(data)

# Export base data to Excel
file_name = "Dynamic_Financial_Report.xlsx"
df.to_excel(file_name, sheet_name="Quarterly Performance", index=False)


# ==========================================
# 2. WRITE NATIVE EXCEL FORMULAS & FORMATTING
# ==========================================
wb = load_workbook(file_name)
ws = wb["Quarterly Performance"]

# Headers for calculated columns in Excel
ws["E1"] = "Total Sales"
ws["F1"] = "Performance Tier"

# Apply bold headers
for col in range(1, 7):
    ws.cell(row=1, column=col).font = Font(bold=True)

# Write dynamic formulas down each row (Rows 2 to 9)
for row in range(2, 10):
    # Excel Formula for Total Sales: =SUM(C2:D2)
    ws[f"E{row}"] = f"=SUM(C{row}:D{row})"

    # Excel Formula for Tier: =IF(E2>=40000, "Top Tier", "Standard")
    ws[f"F{row}"] = f'=IF(E{row}>=40000, "Top Tier", "Standard")'


# ==========================================
# 3. ADD SUMMARY TABLE WITH EXCEL FORMULAS
# (Excel equivalent: Creating a KPI block with SUMIFS / AVERAGEIFS)
# ==========================================
ws["A12"] = "Summary Metric"
ws["B12"] = "Value"
ws["A12"].font = Font(bold=True)
ws["B12"].font = Font(bold=True)

# Inject native Excel summary formulas
ws["A13"] = "Total Company Sales"
ws["B13"] = "=SUM(E2:E9)"

ws["A14"] = "Average Sales Per Rep"
ws["B14"] = "=AVERAGE(E2:E9)"

ws["A15"] = "North Region Total Sales"
ws["B15"] = '=SUMIF(A2:A9, "North", E2:E9)'


# ==========================================
# 4. APPLY COLOR SCALE (HEATMAP) & AUTO-FIT
# ==========================================
# Apply a 3-Color Scale to Total Sales (Column E)
# Green for High, Yellow for Mid, Red for Low
color_scale = ColorScaleRule(
    start_type="min",
    start_color="F8696B",  # Soft Red
    mid_type="percentile",
    mid_value=50,
    mid_color="FFEB84",  # Soft Yellow
    end_type="max",
    end_color="63BE7B",  # Soft Green
)
ws.conditional_formatting.add("E2:E9", color_scale)

# Number Formatting (Currency)
for row in range(2, 10):
    ws[f"C{row}"].number_format = "$#,##0"
    ws[f"D{row}"].number_format = "$#,##0"
    ws[f"E{row}"].number_format = "$#,##0"

ws["B13"].number_format = "$#,##0"
ws["B14"].number_format = "$#,##0"
ws["B15"].number_format = "$#,##0"

# Auto-adjust column widths so no text is cut off (prevents ### display errors)
for col in ws.columns:
    max_len = max(len(str(cell.value or "")) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save the interactive workbook
wb.save(file_name)
print(
    f"Report successfully saved to '{file_name}' with interactive Excel formulas and heatmaps!"
)